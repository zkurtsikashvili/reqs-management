"""
Excel Handler - Inserts requirements into Excel file.
"""

from pathlib import Path
from typing import Dict, Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

DATA_DIR = Path(__file__).parent.parent / "data"
EXCEL_FILE_PATH = DATA_DIR / "requirements_data.xlsx"

COLUMNS = [
    "target_field_name", "description", "target_datamart", "target_field_type", "primary_key",
    "business_key", "nullable", "default_value", "source_field", "source_object_event",
    "source_system_topic", "source_type", "transformation_rules", "is_derived_value",
    "derived_value_logic", "data_quality_rules", "foreign_key_reference_table",
    "foreign_key_reference_field", "pii_flag", "sensitivity_level", "security_rule",
    "retention_policy", "archiving_policy", "source_retention", "source_archiving_policy",
    "data_owner", "data_steward", "comment_notes", "storage", "latency_requirements",
    "is_renamed", "sla_datamart_level", "archiving_datamart_level", "retention_datamart_level"
]


def append_requirement_to_excel(requirement: Dict[str, Any]) -> bool:
    """Append a requirement to Excel file."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    
    try:
        if EXCEL_FILE_PATH.exists():
            workbook = load_workbook(EXCEL_FILE_PATH)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Requirements"
            
            # Create header row
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, col_name in enumerate(COLUMNS, start=1):
                cell = sheet.cell(row=1, column=col_idx)
                cell.value = col_name.replace("_", " ").title()
                cell.fill = header_fill
                cell.font = header_font
                sheet.column_dimensions[cell.column_letter].width = 20
        
        # Add requirement data
        next_row = sheet.max_row + 1
        for col_idx, col_name in enumerate(COLUMNS, start=1):
            sheet.cell(row=next_row, column=col_idx, value=requirement.get(col_name, ""))
        
        workbook.save(EXCEL_FILE_PATH)
        workbook.close()
        return True
    except Exception as e:
        print(f"Error writing to Excel: {e}")
        return False
